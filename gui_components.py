"""
GUI组件模块
包含应用程序的GUI相关组件和对话框
"""

import tkinter as tk
from tkinter import messagebox, filedialog
import customtkinter as ctk
from typing import List, Dict, Any, Callable, Optional, Tuple
import os

# 导入配置模块
from config import AVAILABLE_THEMES


class ThemeSwitcher:
    """主题切换工具类"""
    
    @staticmethod
    def apply_theme(theme_name: str) -> None:
        """应用主题"""
        ctk.set_appearance_mode(theme_name)
    
    @staticmethod
    def toggle_theme(current_theme: str) -> str:
        """切换主题"""
        if current_theme == "dark":
            new_theme = "light"
        else:
            new_theme = "dark"
        
        ThemeSwitcher.apply_theme(new_theme)
        return new_theme


class AboutDialog:
    """关于对话框"""
    
    def __init__(self, parent, app_name: str, app_version: str, core_version: str, title_font):
        """初始化关于对话框"""
        self.dialog = ctk.CTkToplevel(parent)
        self.dialog.title("关于")
        self.dialog.geometry("400x300")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # 创建内容框架
        frame = ctk.CTkFrame(self.dialog)
        frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # 应用名称
        app_name_label = ctk.CTkLabel(frame, text=app_name, font=title_font)
        app_name_label.pack(pady=(0, 5))
        
        # 版本信息
        version_label = ctk.CTkLabel(frame, text=f"版本: {app_version}")
        version_label.pack()
        
        rust_version = ctk.CTkLabel(frame, text=f"核心算法版本: {core_version}")
        rust_version.pack(pady=(0, 10))
        
        # 应用描述
        description = """
这是一个高性能的子集组合求和软件，用于解决从一组数字中找出和为特定目标值的子集问题。
软件采用Rust实现核心算法，Python构建GUI界面，充分利用多线程CPU，追求极致性能。
        """
        desc_label = ctk.CTkLabel(frame, text=description, wraplength=350, justify="center")
        desc_label.pack(pady=10)
        
        # 关闭按钮
        close_button = ctk.CTkButton(frame, text="关闭", command=self.dialog.destroy)
        close_button.pack(pady=10)
        
        # 居中对话框
        self._center_dialog()
    
    def _center_dialog(self):
        """居中对话框"""
        self.dialog.update_idletasks()
        width = self.dialog.winfo_width()
        height = self.dialog.winfo_height()
        x = (self.dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (height // 2)
        self.dialog.geometry('{}x{}+{}+{}'.format(width, height, x, y))


class SettingsDialog:
    """设置对话框"""
    
    def __init__(self, parent, config, save_callback):
        """初始化设置对话框"""
        self.parent = parent
        self.config = config
        self.save_callback = save_callback
        
        self.dialog = ctk.CTkToplevel(parent)
        self.dialog.title("设置")
        self.dialog.geometry("500x400")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # 创建内容框架
        self.main_frame = ctk.CTkFrame(self.dialog)
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # 创建选项卡
        self.tabview = ctk.CTkTabview(self.main_frame)
        self.tabview.pack(fill=tk.BOTH, expand=True)
        
        # 添加选项卡
        self.tab_general = self.tabview.add("常规")
        self.tab_appearance = self.tabview.add("外观")
        self.tab_files = self.tabview.add("文件")
        
        # 初始化各个选项卡
        self._setup_general_tab()
        self._setup_appearance_tab()
        self._setup_files_tab()
        self._setup_buttons()
        
        # 居中对话框
        self._center_dialog()
    
    def _setup_general_tab(self):
        """设置常规选项卡"""
        # 显示可视化选项
        self.show_viz_var = tk.BooleanVar(value=self.config.get("show_visualization", True))
        show_viz_check = ctk.CTkCheckBox(self.tab_general, text="显示结果可视化", variable=self.show_viz_var)
        show_viz_check.pack(anchor=tk.W, pady=(10, 5))
        
        # 自动保存结果选项
        self.auto_save_var = tk.BooleanVar(value=self.config.get("auto_save_results", False))
        auto_save_check = ctk.CTkCheckBox(self.tab_general, text="计算完成后自动保存结果", variable=self.auto_save_var)
        auto_save_check.pack(anchor=tk.W, pady=(0, 5))
        
        # 内存限制默认值
        memory_frame = ctk.CTkFrame(self.tab_general)
        memory_frame.pack(fill=tk.X, pady=(10, 5))
        
        memory_label = ctk.CTkLabel(memory_frame, text="默认内存限制 (MB):")
        memory_label.pack(side=tk.LEFT)
        
        self.memory_var = tk.StringVar(value=str(self.config.get("default_memory_limit", 500)))
        memory_entry = ctk.CTkEntry(memory_frame, textvariable=self.memory_var, width=100)
        memory_entry.pack(side=tk.LEFT, padx=(5, 0))
        
        # 最大解决方案默认值
        solutions_frame = ctk.CTkFrame(self.tab_general)
        solutions_frame.pack(fill=tk.X, pady=(10, 5))
        
        solutions_label = ctk.CTkLabel(solutions_frame, text="默认解决方案数量:")
        solutions_label.pack(side=tk.LEFT)
        
        self.solutions_var = tk.StringVar(value=str(self.config.get("default_max_solutions", 1)))
        solutions_entry = ctk.CTkEntry(solutions_frame, textvariable=self.solutions_var, width=100)
        solutions_entry.pack(side=tk.LEFT, padx=(5, 0))
    
    def _setup_appearance_tab(self):
        """设置外观选项卡"""
        theme_frame = ctk.CTkFrame(self.tab_appearance)
        theme_frame.pack(fill=tk.X, pady=(10, 5))
        
        theme_label = ctk.CTkLabel(theme_frame, text="主题模式:")
        theme_label.pack(side=tk.LEFT)
        
        self.theme_var = tk.StringVar(value=self.config.get("theme", "system"))
        theme_option = ctk.CTkOptionMenu(
            theme_frame, 
            values=AVAILABLE_THEMES,
            variable=self.theme_var
        )
        theme_option.pack(side=tk.LEFT, padx=(5, 0), fill=tk.X, expand=True)
        
        # 预览按钮
        preview_button = ctk.CTkButton(
            self.tab_appearance, 
            text="预览主题", 
            command=lambda: ThemeSwitcher.apply_theme(self.theme_var.get())
        )
        preview_button.pack(anchor=tk.W, pady=(10, 5))
    
    def _setup_files_tab(self):
        """设置文件选项卡"""
        recent_label = ctk.CTkLabel(self.tab_files, text="最近使用的文件:")
        recent_label.pack(anchor=tk.W, pady=(10, 5))
        
        # 最近文件列表框
        recent_frame = ctk.CTkFrame(self.tab_files)
        recent_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # 文件列表显示
        recent_files = self.config.get("recent_files", [])
        self.recent_text = ctk.CTkTextbox(recent_frame, height=150)
        self.recent_text.pack(fill=tk.BOTH, expand=True)
        
        for file_path in recent_files:
            self.recent_text.insert(tk.END, f"{file_path}\n")
        self.recent_text.configure(state="disabled")
        
        # 清除按钮
        clear_button = ctk.CTkButton(
            self.tab_files, 
            text="清除最近文件列表", 
            command=self._clear_recent_files
        )
        clear_button.pack(anchor=tk.W, pady=(5, 0))
    
    def _setup_buttons(self):
        """设置按钮区域"""
        button_frame = ctk.CTkFrame(self.main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        # 确定按钮
        ok_button = ctk.CTkButton(
            button_frame, 
            text="确定", 
            command=self._save_settings
        )
        ok_button.pack(side=tk.RIGHT, padx=(5, 0))
        
        # 取消按钮
        cancel_button = ctk.CTkButton(
            button_frame, 
            text="取消", 
            command=self.dialog.destroy
        )
        cancel_button.pack(side=tk.RIGHT, padx=(5, 0))
    
    def _save_settings(self):
        """保存设置"""
        try:
            # 验证数字设置
            try:
                memory = int(self.memory_var.get())
                if memory < 100:
                    raise ValueError("内存限制不能小于100MB")
                
                solutions = int(self.solutions_var.get())
                if solutions < 1:
                    raise ValueError("解决方案数量必须大于0")
                
            except ValueError as e:
                messagebox.showerror("参数错误", str(e))
                return
            
            # 调用回调函数
            self.save_callback(
                self.show_viz_var.get(),
                self.auto_save_var.get(),
                self.memory_var.get(),
                self.solutions_var.get(),
                self.theme_var.get(),
                self.dialog
            )
                
        except Exception as e:
            messagebox.showerror("保存设置错误", f"保存设置时出错: {str(e)}")
    
    def _clear_recent_files(self):
        """清除最近使用的文件列表"""
        self.config.clear_recent_files()
        self.config.save_config()
        self.recent_text.configure(state="normal")
        self.recent_text.delete(1.0, tk.END)
        self.recent_text.configure(state="disabled")
    
    def _center_dialog(self):
        """居中对话框"""
        self.dialog.update_idletasks()
        width = self.dialog.winfo_width()
        height = self.dialog.winfo_height()
        x = (self.dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (height // 2)
        self.dialog.geometry('{}x{}+{}+{}'.format(width, height, x, y))


class ExcelImportDialog:
    """Excel导入对话框"""
    
    def __init__(self, parent, file_path, import_callback):
        """初始化Excel导入对话框"""
        import openpyxl
        
        self.parent = parent
        self.import_callback = import_callback
        
        # 使用openpyxl加载Excel文件
        self.workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet_names = self.workbook.sheetnames
        
        # 创建对话框
        self.dialog = ctk.CTkToplevel(parent)
        self.dialog.title("从Excel导入")
        self.dialog.geometry("400x400")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # 创建内容框架
        self.frame = ctk.CTkFrame(self.dialog)
        self.frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 工作表选择
        sheet_frame = ctk.CTkFrame(self.frame)
        sheet_frame.pack(fill=tk.X, pady=(0, 10))
        
        sheet_label = ctk.CTkLabel(sheet_frame, text="选择工作表:")
        sheet_label.pack(side=tk.LEFT)
        
        self.sheet_var = tk.StringVar(value=sheet_names[0])
        sheet_combo = ctk.CTkOptionMenu(sheet_frame, values=sheet_names, variable=self.sheet_var)
        sheet_combo.pack(side=tk.LEFT, padx=(5, 0), fill=tk.X, expand=True)
        
        # 列选择
        column_frame = ctk.CTkFrame(self.frame)
        column_frame.pack(fill=tk.X, pady=(0, 10))
        
        column_label = ctk.CTkLabel(column_frame, text="选择数据列:")
        column_label.pack(side=tk.LEFT)
        
        self.column_var = tk.StringVar(value="A")
        column_entry = ctk.CTkEntry(column_frame, textvariable=self.column_var)
        column_entry.pack(side=tk.LEFT, padx=(5, 0), fill=tk.X, expand=True)
        
        # 行范围选择
        row_frame = ctk.CTkFrame(self.frame)
        row_frame.pack(fill=tk.X, pady=(0, 10))
        
        start_row_label = ctk.CTkLabel(row_frame, text="起始行:")
        start_row_label.pack(side=tk.LEFT)
        
        self.start_row_var = tk.StringVar(value="1")
        start_row_entry = ctk.CTkEntry(row_frame, textvariable=self.start_row_var, width=50)
        start_row_entry.pack(side=tk.LEFT, padx=(5, 10))
        
        end_row_label = ctk.CTkLabel(row_frame, text="结束行:")
        end_row_label.pack(side=tk.LEFT)
        
        self.end_row_var = tk.StringVar(value="100")
        end_row_entry = ctk.CTkEntry(row_frame, textvariable=self.end_row_var, width=50)
        end_row_entry.pack(side=tk.LEFT, padx=(5, 0))
        
        # 预览按钮
        preview_button = ctk.CTkButton(
            self.frame, 
            text="预览数据", 
            command=self._preview_data
        )
        preview_button.pack(anchor=tk.W, pady=(5, 10))
        
        # 预览文本区域
        self.preview_text = ctk.CTkTextbox(self.frame, height=150)
        self.preview_text.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # 按钮区域
        button_frame = ctk.CTkFrame(self.frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        # 导入按钮
        import_button = ctk.CTkButton(
            button_frame, 
            text="导入", 
            command=self._import_data
        )
        import_button.pack(side=tk.LEFT, padx=(0, 5))
        
        # 取消按钮
        cancel_button = ctk.CTkButton(button_frame, text="取消", command=self.dialog.destroy)
        cancel_button.pack(side=tk.LEFT)
        
        # 居中对话框
        self._center_dialog()
    
    def _preview_data(self):
        """预览Excel数据"""
        try:
            # 获取工作表
            sheet = self.workbook[self.sheet_var.get()]
            
            # 解析参数
            try:
                start = int(self.start_row_var.get())
                end = int(self.end_row_var.get())
                if start < 1:
                    start = 1
                if end < start:
                    end = start
            except ValueError:
                messagebox.showerror("参数错误", "行号必须是整数")
                return
            
            # 清空预览
            self.preview_text.configure(state="normal")
            self.preview_text.delete(1.0, tk.END)
            
            # 提取数据并显示前10行
            count = 0
            values = []
            
            for row_idx in range(start, min(end + 1, sheet.max_row + 1)):
                cell = sheet[f"{self.column_var.get().upper()}{row_idx}"]
                if cell.value is not None:
                    try:
                        value = float(cell.value)
                        if value <= 0:
                            continue
                        values.append(value)
                        count += 1
                        if count <= 10:  # 只预览前10行
                            self.preview_text.insert(tk.END, f"{value:.2f}\n")
                    except (ValueError, TypeError):
                        pass
            
            # 添加摘要信息
            if count > 10:
                self.preview_text.insert(tk.END, f"\n... 共找到 {count} 个有效数字")
            elif count == 0:
                self.preview_text.insert(tk.END, "未找到有效数字")
                
            self.preview_text.configure(state="disabled")
                
        except Exception as e:
            self.preview_text.configure(state="normal")
            self.preview_text.delete(1.0, tk.END)
            self.preview_text.insert(tk.END, f"预览出错: {str(e)}")
            self.preview_text.configure(state="disabled")
    
    def _import_data(self):
        """导入Excel数据"""
        # 调用回调函数
        self.import_callback(
            self.workbook, 
            self.sheet_var.get(), 
            self.column_var.get(),
            self.start_row_var.get(), 
            self.end_row_var.get(), 
            self.dialog
        )
    
    def _center_dialog(self):
        """居中对话框"""
        self.dialog.update_idletasks()
        width = self.dialog.winfo_width()
        height = self.dialog.winfo_height()
        x = (self.dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (height // 2)
        self.dialog.geometry('{}x{}+{}+{}'.format(width, height, x, y))


class ExcelImportDialog:
    """Excel导入对话框"""
    
    def __init__(self, parent, file_path, import_callback):
        """初始化Excel导入对话框"""
        import openpyxl
        
        self.parent = parent
        self.import_callback = import_callback
        
        # 使用openpyxl加载Excel文件
        self.workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet_names = self.workbook.sheetnames
        
        # 创建对话框
        self.dialog = ctk.CTkToplevel(parent)
        self.dialog.title("从Excel导入")
        self.dialog.geometry("400x400")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # 创建内容框架
        self.frame = ctk.CTkFrame(self.dialog)
        self.frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 工作表选择
        sheet_frame = ctk.CTkFrame(self.frame)
        sheet_frame.pack(fill=tk.X, pady=(0, 10))
        
        sheet_label = ctk.CTkLabel(sheet_frame, text="选择工作表:")
        sheet_label.pack(side=tk.LEFT)
        
        self.sheet_var = tk.StringVar(value=sheet_names[0])
        sheet_combo = ctk.CTkOptionMenu(sheet_frame, values=sheet_names, variable=self.sheet_var)
        sheet_combo.pack(side=tk.LEFT, padx=(5, 0), fill=tk.X, expand=True)
        
        # 列选择
        column_frame = ctk.CTkFrame(self.frame)
        column_frame.pack(fill=tk.X, pady=(0, 10))
        
        column_label = ctk.CTkLabel(column_frame, text="选择数据列:")
        column_label.pack(side=tk.LEFT)
        
        self.column_var = tk.StringVar(value="A")
        column_entry = ctk.CTkEntry(column_frame, textvariable=self.column_var)
        column_entry.pack(side=tk.LEFT, padx=(5, 0), fill=tk.X, expand=True)
        
        # 行范围选择
        row_frame = ctk.CTkFrame(self.frame)
        row_frame.pack(fill=tk.X, pady=(0, 10))
        
        start_row_label = ctk.CTkLabel(row_frame, text="起始行:")
        start_row_label.pack(side=tk.LEFT)
        
        self.start_row_var = tk.StringVar(value="1")
        start_row_entry = ctk.CTkEntry(row_frame, textvariable=self.start_row_var, width=50)
        start_row_entry.pack(side=tk.LEFT, padx=(5, 10))
        
        end_row_label = ctk.CTkLabel(row_frame, text="结束行:")
        end_row_label.pack(side=tk.LEFT)
        
        self.end_row_var = tk.StringVar(value="100")
        end_row_entry = ctk.CTkEntry(row_frame, textvariable=self.end_row_var, width=50)
        end_row_entry.pack(side=tk.LEFT, padx=(5, 0))
        
        # 预览按钮
        preview_button = ctk.CTkButton(
            self.frame, 
            text="预览数据", 
            command=self._preview_data
        )
        preview_button.pack(anchor=tk.W, pady=(5, 10))
        
        # 预览文本区域
        self.preview_text = ctk.CTkTextbox(self.frame, height=150)
        self.preview_text.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # 按钮区域
        button_frame = ctk.CTkFrame(self.frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        # 导入按钮
        import_button = ctk.CTkButton(
            button_frame, 
            text="导入", 
            command=self._import_data
        )
        import_button.pack(side=tk.LEFT, padx=(0, 5))
        
        # 取消按钮
        cancel_button = ctk.CTkButton(button_frame, text="取消", command=self.dialog.destroy)
        cancel_button.pack(side=tk.LEFT)
        
        # 居中对话框
        self._center_dialog()
    
    def _preview_data(self):
        """预览Excel数据"""
        try:
            # 获取工作表
            sheet = self.workbook[self.sheet_var.get()]
            
            # 解析参数
            try:
                start = int(self.start_row_var.get())
                end = int(self.end_row_var.get())
                if start < 1:
                    start = 1
                if end < start:
                    end = start
            except ValueError:
                messagebox.showerror("参数错误", "行号必须是整数")
                return
            
            # 清空预览
            self.preview_text.configure(state="normal")
            self.preview_text.delete(1.0, tk.END)
            
            # 提取数据并显示前10行
            count = 0
            values = []
            
            for row_idx in range(start, min(end + 1, sheet.max_row + 1)):
                cell = sheet[f"{self.column_var.get().upper()}{row_idx}"]
                if cell.value is not None:
                    try:
                        value = float(cell.value)
                        if value <= 0:
                            continue
                        values.append(value)
                        count += 1
                        if count <= 10:  # 只预览前10行
                            self.preview_text.insert(tk.END, f"{value:.2f}\n")
                    except (ValueError, TypeError):
                        pass
            
            # 添加摘要信息
            if count > 10:
                self.preview_text.insert(tk.END, f"\n... 共找到 {count} 个有效数字")
            elif count == 0:
                self.preview_text.insert(tk.END, "未找到有效数字")
                
            self.preview_text.configure(state="disabled")
                
        except Exception as e:
            self.preview_text.configure(state="normal")
            self.preview_text.delete(1.0, tk.END)
            self.preview_text.insert(tk.END, f"预览出错: {str(e)}")
            self.preview_text.configure(state="disabled")
    
    def _import_data(self):
        """导入Excel数据"""
        # 调用回调函数
        self.import_callback(
            self.workbook, 
            self.sheet_var.get(), 
            self.column_var.get(),
            self.start_row_var.get(), 
            self.end_row_var.get(), 
            self.dialog
        )
    
    def _center_dialog(self):
        """居中对话框"""
        self.dialog.update_idletasks()
        width = self.dialog.winfo_width()
        height = self.dialog.winfo_height()
        x = (self.dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (height // 2)
        self.dialog.geometry('{}x{}+{}+{}'.format(width, height, x, y))


# 新增的UI模块化函数
def create_input_area(parent_frame, subtitle_font, config, load_callback, save_callback, start_callback, stop_callback):
    """创建输入区域
    
    Args:
        parent_frame: 父框架
        subtitle_font: 副标题字体
        config: 配置对象
        load_callback: 加载文件回调
        save_callback: 保存文件回调
        start_callback: 开始计算回调
        stop_callback: 停止计算回调
        
    Returns:
        (numbers_text, load_button, save_button, target_entry, solutions_var, 
        solutions_spinbox, memory_var, memory_spinbox, start_button, stop_button)
    """
    input_frame = ctk.CTkFrame(parent_frame)
    input_frame.pack(fill=tk.BOTH, expand=True)
    
    input_label = ctk.CTkLabel(input_frame, text="输入数据", font=subtitle_font)
    input_label.pack(anchor=tk.W, padx=10, pady=5)
    
    # 数字输入文本框
    numbers_frame = ctk.CTkFrame(input_frame)
    numbers_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
    
    numbers_label = ctk.CTkLabel(numbers_frame, text="输入数字（每行一个，支持最多两位小数的正数）:")
    numbers_label.pack(anchor=tk.W, pady=(5, 0))
    
    numbers_text = ctk.CTkTextbox(numbers_frame, width=400, height=200)
    numbers_text.pack(fill=tk.BOTH, expand=True, pady=5)
    
    # 文件操作按钮
    file_frame = ctk.CTkFrame(input_frame)
    file_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
    
    load_button = ctk.CTkButton(file_frame, text="从文件加载", command=load_callback)
    load_button.pack(side=tk.LEFT, padx=(0, 5))
    
    save_button = ctk.CTkButton(file_frame, text="保存到文件", command=save_callback)
    save_button.pack(side=tk.LEFT)
    
    # 使用简单布局实现自适应的参数设置区域
    params_frame = ctk.CTkFrame(input_frame)
    params_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
    
    # 目标和
    target_frame = ctk.CTkFrame(params_frame)
    target_frame.pack(fill=tk.X, pady=5)
    
    target_label = ctk.CTkLabel(target_frame, text="目标和:")
    target_label.pack(side=tk.LEFT)
    
    target_entry = ctk.CTkEntry(target_frame)
    target_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))
    
    # 解决方案数量
    solutions_frame = ctk.CTkFrame(params_frame)
    solutions_frame.pack(fill=tk.X, pady=5)
    
    solutions_label = ctk.CTkLabel(solutions_frame, text="最大解决方案数量:")
    solutions_label.pack(side=tk.LEFT)
    
    solutions_var = tk.StringVar(value=config.get("default_max_solutions"))
    solutions_spinbox = ctk.CTkEntry(solutions_frame, textvariable=solutions_var, width=80)
    solutions_spinbox.pack(side=tk.LEFT, padx=(5, 0))
    
    # 内存限制
    memory_frame = ctk.CTkFrame(params_frame)
    memory_frame.pack(fill=tk.X, pady=5)
    
    memory_label = ctk.CTkLabel(memory_frame, text="内存限制 (MB):")
    memory_label.pack(side=tk.LEFT)
    
    memory_var = tk.StringVar(value=config.get("default_memory_limit"))
    memory_spinbox = ctk.CTkEntry(memory_frame, textvariable=memory_var, width=80)
    memory_spinbox.pack(side=tk.LEFT, padx=(5, 0))
    
    # 操作按钮布局
    button_frame = ctk.CTkFrame(input_frame)
    button_frame.pack(fill=tk.X, padx=10, pady=(10, 10))
    
    start_button = ctk.CTkButton(button_frame, text="开始计算", command=start_callback, 
                                fg_color="#4CAF50", hover_color="#3e8e41")
    start_button.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
    
    stop_button = ctk.CTkButton(button_frame, text="停止", command=stop_callback, 
                              state="disabled", fg_color="#F44336", hover_color="#d32f2f")
    stop_button.pack(side=tk.LEFT, fill=tk.X, expand=True)
    
    return (numbers_text, load_button, save_button, target_entry, solutions_var, 
            solutions_spinbox, memory_var, memory_spinbox, start_button, stop_button)


def create_result_area(parent_frame, subtitle_font, export_callback):
    """创建结果区域
    
    Args:
        parent_frame: 父框架
        subtitle_font: 副标题字体
        export_callback: 导出回调
        
    Returns:
        (progress_bar, progress_label, result_text, export_button, canvas_frame)
    """
    result_frame = ctk.CTkFrame(parent_frame)
    result_frame.pack(fill=tk.BOTH, expand=True)
    
    result_label = ctk.CTkLabel(result_frame, text="计算结果", font=subtitle_font)
    result_label.pack(anchor=tk.W, padx=10, pady=5)
    
    # 进度条
    progress_frame = ctk.CTkFrame(result_frame)
    progress_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
    
    progress_label_text = ctk.CTkLabel(progress_frame, text="计算进度:")
    progress_label_text.pack(anchor=tk.W)
    
    progress_bar = ctk.CTkProgressBar(progress_frame)
    progress_bar.pack(fill=tk.X, pady=(5, 0))
    progress_bar.set(0)
    
    # 进度文本
    progress_label = ctk.CTkLabel(progress_frame, text="0.0%")
    progress_label.pack(anchor=tk.E, pady=(5, 0))
    
    # 结果文本区
    result_header_frame = ctk.CTkFrame(result_frame)
    result_header_frame.pack(fill=tk.X, padx=10, pady=(10, 5))
    
    result_title = ctk.CTkLabel(result_header_frame, text="找到的子集:")
    result_title.pack(side=tk.LEFT)
    
    # 导出结果按钮
    export_button = ctk.CTkButton(result_header_frame, text="导出结果", 
                                 command=export_callback, state="disabled")
    export_button.pack(side=tk.RIGHT)
    
    result_text = ctk.CTkTextbox(result_frame, width=400, height=200)
    result_text.pack(fill=tk.BOTH, expand=True, padx=10)
    
    # 可视化区域
    viz_frame = ctk.CTkFrame(result_frame)
    viz_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(10, 10))
    
    viz_label = ctk.CTkLabel(viz_frame, text="结果可视化", font=subtitle_font)
    viz_label.pack(anchor=tk.W, pady=5)
    
    canvas_frame = ctk.CTkFrame(viz_frame)
    canvas_frame.pack(fill=tk.BOTH, expand=True, pady=(5, 5))
    
    return (progress_bar, progress_label, result_text, export_button, canvas_frame)


def setup_menu(root, load_callback, save_callback, copy_callback, paste_callback, 
              theme_callback, settings_callback, about_callback):
    """设置菜单
    
    Args:
        root: 根窗口
        load_callback: 加载文件回调
        save_callback: 保存文件回调
        copy_callback: 复制回调
        paste_callback: 粘贴回调
        theme_callback: 主题切换回调
        settings_callback: 设置回调
        about_callback: 关于回调
    """
    menubar = tk.Menu(root)
    root.config(menu=menubar)
    
    # 文件菜单
    file_menu = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="文件", menu=file_menu)
    file_menu.add_command(label="打开", command=load_callback)
    file_menu.add_command(label="保存", command=save_callback)
    file_menu.add_separator()
    file_menu.add_command(label="退出", command=root.quit)
    
    # 编辑菜单
    edit_menu = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="编辑", menu=edit_menu)
    edit_menu.add_command(label="复制", command=copy_callback)
    edit_menu.add_command(label="粘贴", command=paste_callback)
    
    # 视图菜单
    view_menu = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="视图", menu=view_menu)
    view_menu.add_command(label="切换主题", command=theme_callback)
    view_menu.add_command(label="设置", command=settings_callback)
    
    # 帮助菜单
    help_menu = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="帮助", menu=help_menu)
    help_menu.add_command(label="关于", command=about_callback)


def handle_file_load(root, numbers_text, status_var, config, excel_callback):
    """处理文件加载
    
    Args:
        root: 根窗口
        numbers_text: 数字文本框
        status_var: 状态变量
        config: 配置对象
        excel_callback: Excel加载回调
    """
    from file_operations import load_text_file
    
    file_types = [
        ("文本文件", "*.txt"), 
        ("Excel文件", "*.xlsx;*.xls"), 
        ("所有文件", "*.*")
    ]
    
    file_path = tk.filedialog.askopenfilename(
        title="选择数字文件",
        filetypes=file_types
    )
    
    if not file_path:
        return
    
    try:
        if file_path.lower().endswith(('.xlsx', '.xls')):
            excel_callback(file_path)
        else:
            # 文本文件处理
            content, valid_numbers = load_text_file(file_path)
            
            numbers_text.delete(1.0, tk.END)
            numbers_text.insert(tk.END, content)
            status_var.set(f"从文件加载了 {len(valid_numbers)} 个有效数字: {os.path.basename(file_path)}")
        
        # 添加到最近文件列表
        config.add_recent_file(file_path)
        config.save_config()
        
    except Exception as e:
        tk.messagebox.showerror("加载错误", f"无法加载文件: {str(e)}")


def handle_file_save(root, numbers_text, status_var, config):
    """处理文件保存
    
    Args:
        root: 根窗口
        numbers_text: 数字文本框
        status_var: 状态变量
        config: 配置对象
    """
    from file_operations import save_text_file
    
    file_path = tk.filedialog.asksaveasfilename(
        title="保存数字文件",
        defaultextension=".txt",
        filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")]
    )
    
    if not file_path:
        return
    
    try:
        content = numbers_text.get(1.0, tk.END)
        if save_text_file(content, file_path):
            # 添加到最近文件列表
            config.add_recent_file(file_path)
            config.save_config()
            
            status_var.set(f"数据已保存到文件: {os.path.basename(file_path)}")
    except Exception as e:
        tk.messagebox.showerror("保存错误", f"无法保存文件: {str(e)}")
